#!/usr/bin/env python3
"""
curate.py — the on-demand "refresh" for the living brand-first board.

Run it whenever you want a fresh board:

    .venv/bin/python curate.py                 # refresh the board
    .venv/bin/python curate.py --validate-boards   # ping every board, report counts

Flow (round-trip preserve — your manual edits survive):
  1. abort if the xlsx is open in Excel (~$ lock)
  2. load curated_postings.json (the store)
  3. read your manual edits back out of the existing board (keyed by hidden _id)
     and merge them into the store
  4. one-time: seed My Applications from the old Application Tracker history
  5. pull every target company's board (lane-1 brand-first), full JDs via the
     ATS JSON APIs; score each by hotness; upsert into the store
  6. re-score every stored posting (recency decays daily); brand-board postings
     that fell off their board for 2 runs -> stale (drop from queue, keep in JSON)
  7. atomically write the store + regenerate Curated Board.xlsx

Runs locally (no 120s wall). The existing Application Tracker.xlsx is never touched.
"""
from __future__ import annotations

import argparse
import asyncio
import socket
import sys
import time
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path


def _wait_for_network(hosts=("boards-api.greenhouse.io", "api.anthropic.com", "api.telegram.org"),
                      tries=18, delay=5) -> bool:
    """Scheduled runs fire as the Mac wakes, and the network settles UNEVENLY — the job
    boards resolve before OpenAI/Telegram do. Wait (up to ~90s) until ALL critical hosts
    resolve, so a wake-run doesn't harvest fine but then DNS-fail the fit pass + Telegram
    notify (observed 2026-06-24: 94 unscored + no Telegram on a 9:36 wake-run). If a host
    never comes up, proceed anyway — harvest may still work, fit/notify degrade gracefully,
    and the harvest-collapse guard protects the board."""
    for i in range(tries):
        try:
            for h in hosts:
                socket.gethostbyname(h)
            return True
        except OSError:
            if i == 0:
                print("[refresh] network still settling — waiting for full connectivity…",
                      file=sys.stderr)
            time.sleep(delay)
    print("[refresh] some hosts still unresolved after wait — proceeding (fit/notify may degrade)",
          file=sys.stderr)
    return False

# Telegram (reuse the Hermes bot — same chat the other crons send to)
TELEGRAM_CHAT_ID = "696500863"
HERMES_ENV = Path.home() / ".hermes" / ".env"


def _hermes_env(key: str) -> str | None:
    if HERMES_ENV.exists():
        for line in HERMES_ENV.read_text(encoding="utf-8", errors="replace").splitlines():
            line = line.strip()
            if line.startswith(key + "="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    import os
    return os.environ.get(key)


def _send_telegram(text: str) -> bool:
    token = _hermes_env("TELEGRAM_BOT_TOKEN")
    if not token:
        print("[notify] no TELEGRAM_BOT_TOKEN in ~/.hermes/.env — skipped", file=sys.stderr)
        return False
    try:
        data = urllib.parse.urlencode({
            "chat_id": TELEGRAM_CHAT_ID, "text": text,
            "parse_mode": "Markdown", "disable_web_page_preview": "true",
        }).encode()
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendMessage", data=data)
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.status == 200
    except Exception as e:  # noqa: BLE001
        print(f"[notify] telegram send failed: {e}", file=sys.stderr)
        return False


def _new_postings_digest(store, new_cids: set) -> str | None:
    rows = []
    for cid in new_cids:
        m = (store.get(cid) or {}).get("machine", {})
        if m.get("dead") or not m.get("company"):
            continue
        rows.append(m)
    if not rows:
        return None
    # clean (applyable) vs disqualified — surface both, separated
    def _disq(m):
        return (m.get("fit_disqualifier") or "none") not in ("none", "", None)
    clean = sorted([m for m in rows if not _disq(m)], key=lambda m: -int(m.get("hotness", 0) or 0))
    flagged = [m for m in rows if _disq(m)]
    n = len(rows)

    def _fmt(m):
        fit = m.get("fit_score")
        fitstr = f" · fit {fit}" if isinstance(fit, (int, float)) else ""
        return (f"• [{m.get('tier', '?')}] *{m.get('company', '?')}* — "
                f"{str(m.get('role', '?'))[:46]}{fitstr} {m.get('fresh', '')}".rstrip())

    # CYCLE-OPEN WATCHER: a new tier-S/A role means a top brand just opened (or
    # added) an intern req. These are rare + time-sensitive (top spots fill fast),
    # so lead with them so they never get buried under B-tier volume.
    elite = [m for m in clean if str(m.get("tier", "")).upper() in ("S", "A")]
    rest = [m for m in clean if str(m.get("tier", "")).upper() not in ("S", "A")]

    lines = [f"🔥 *{n} new internship role{'s' if n != 1 else ''}* on your Curated Board:"]
    if elite:
        lines.append(f"\n🚨 *{len(elite)} TOP-BRAND role{'s' if len(elite) != 1 else ''} "
                     f"just opened* — apply early, these fill fast:")
        lines += [_fmt(m) for m in elite[:8]]
        if len(elite) > 8:
            lines.append(f"…and {len(elite) - 8} more top-brand.")
    if rest:
        if elite:
            lines.append("\n*Other new roles:*")
        lines += [_fmt(m) for m in rest[:8]]
        if len(rest) > 8:
            lines.append(f"…and {len(rest) - 8} more.")
    if flagged:
        lines.append(f"\n⚠️ {len(flagged)} flagged (sunk, AI found a disqualifier):")
        for m in flagged[:4]:
            lines.append(f"  • {m.get('company', '?')} — {str(m.get('role', '?'))[:34]}: "
                         f"_{m.get('fit_disqualifier')}_")
    lines.append("\nOpen *Curated Board.xlsx* → 🔥 Curated Queue to apply.")
    return "\n".join(lines)

# Paths come from store_paths.py — the ONE resolver, shared by every script here.
# 🔴 Do NOT reintroduce a local default. This file used to default the store to the
# vault copy, which on the VPS is a stale historical file; a bare `python curate.py`
# then loaded it, orphaned ~700 unrecognised Sheet rows into identity-less entries,
# and rendered them as blank rows. See store_paths.py for the full account.
import os  # noqa: E402
import re  # noqa: E402
sys.path.insert(0, str(Path(__file__).resolve().parent))
import store_paths  # noqa: E402
VAULT = store_paths.vault_root()
STORE_PATH = store_paths.store_path()
XLSX_PATH = store_paths.xlsx_path()
OLD_TRACKER = VAULT / "06 - Internships" / "Job Search" / "Legacy" / "Application Tracker.xlsx"

# Google Sheets mirror — the phone-accessible copy, and during the transition the
# SECOND artifact rather than a replacement (the xlsx stays the fallback and the
# diffable one). CURATED_GSHEET=0 turns it off; CURATED_GSHEET_ID points elsewhere.
GSHEET_ID = os.environ.get("CURATED_GSHEET_ID",
                           "1Kkle7QoKsBMXihoslWjIoMDqKFznwxxA4Y_OgiqJpWI")
GSHEET_ON = os.environ.get("CURATED_GSHEET", "1").lower() not in ("0", "false", "no")

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(VAULT / "Scripts"))

import brand_first_source  # noqa: E402
import wide_net_source  # noqa: E402
from company_boards import boards as _boards  # noqa: E402
from build_curated_xlsx import (REVIEWED_STATUSES, classify_row, is_locked,  # noqa: E402
                                write_board)
try:
    import build_curated_gsheet as gsheet  # noqa: E402
except Exception as _e:                    # noqa: BLE001 - never fail the refresh on import
    gsheet = None
    print(f"[refresh] Google Sheets mirror unavailable: {_e}", file=sys.stderr)
from curated_store import CuratedStore  # noqa: E402
from hotness import hotness, role_lane  # noqa: E402
from internship_scraper import canonical_id, normalize_company_name  # noqa: E402

STALE_STRIKES = 2
# Wide-net sources are ROLLING WINDOWS — falling off a GitHub README is not
# evidence a req closed (measured 2026-08-18: 54% of dead wide-net rows were
# still open on the employer's own ATS). Needs many more strikes than a direct
# API dropout before we believe it.
WIDE_STALE_STRIKES = 14          # ~1 week at 2 runs/day
# A healthy lane-1 harvest is ~68 roles (median over 108 logged runs, min 25).
# Below this, treat lane 1 as collapsed and exempt its postings from striking.
LANE1_MIN_HEALTHY = 25
_ACTIONED = {"applied", "oa", "phone screen", "onsite", "offer", "rejected",
             "networking", "on hold"}

import re as _re  # noqa: E402
_EMOJI = _re.compile("[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF←-⇿⬀-⯿]")


def _clean_company(name: str) -> str:
    """Strip aggregator emoji/symbols (e.g. '🔥Tesla' -> 'Tesla') and tidy."""
    if not name:
        return name
    s = _EMOJI.sub("", name)
    s = _re.sub(r"^[^\w(]+", "", s)            # leading symbols
    return _re.sub(r"\s+", " ", s).strip()


def _score_into(store, cid, base: dict, *, first_seen: str, source: str):
    """Compute hotness for a posting and upsert its machine fields."""
    h = hotness(base["company"], base["role"], base.get("age_days"),
                lane=role_lane(base["role"]))
    existing = (store.get(cid) or {}).get("machine", {})
    store.upsert_machine(cid, {
        **base,
        "source": source,
        "lane": h["lane"], "tier": h["tier"], "brand": h["brand"],
        "role_score": h["role"], "recency": h["recency"],
        "hotness": h["hotness"], "fresh": h["fresh"],
        "dead": False, "fail_count": 0,
        "first_seen": existing.get("first_seen") or first_seen,
        "last_seen": first_seen,
    })


def import_old_tracker_history(store) -> int:
    """One-time, idempotent: pull the actioned rows (Offer/Rejected/Applied/…) from
    the old Application Tracker.xlsx into My Applications so history isn't lost.
    Skips 'To Apply' rows (brand-first rediscovers live openings fresh)."""
    if not OLD_TRACKER.exists():
        return 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(OLD_TRACKER, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return 0
    seeded = 0
    try:
        ws = wb["Apply Board"] if "Apply Board" in wb.sheetnames else wb.worksheets[0]
        header = None
        col = {}
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if header is None:
                if "Company" in cells and "Status" in cells:
                    header = cells
                    col = {h.lower(): i for i, h in enumerate(header)}
                continue

            def g(name):
                i = col.get(name.lower())
                return cells[i] if i is not None and i < len(cells) else ""
            company, role, status = g("company"), g("role"), g("status")
            if not company or not role:
                continue
            if status.strip().lower() not in _ACTIONED:
                continue
            url = g("apply") if g("apply").startswith("http") else ""
            # the old "Apply" cell is just the label "Apply ↗"; real url isn't stored
            # readably, so synthesize a stable history id from company+role
            cid = canonical_id(url) if url.startswith("http") else \
                f"hist/{normalize_company_name(company)}/{role.lower()[:40]}"
            if cid in store.postings:
                continue
            store.upsert_machine(cid, {
                "company": company, "role": role, "location": g("location"),
                "url": url, "ats_type": "manual", "source": "history",
                "cycle": g("cycle"), "lane": g("lane") or "Other",
                "posted_date": "", "age_days": None, "full_jd": "",
                "tier": "", "hotness": 0, "fresh": "", "dead": False,
            })
            store.set_human(cid, {"status": status, "applied_date": g("applied"),
                                  "notes": g("notes")})
            seeded += 1
    finally:
        wb.close()
    return seeded


# 🔴 An `_id` on the board that the store does not know is a PATHOLOGICAL signal, not a
# routine one. The board is rendered FROM the store, and nothing ever deletes a posting
# (entries are flagged `dead`, never removed), so in healthy operation every id read back
# is already present. A handful can appear legitimately — a row hand-typed onto the
# Sheet, or an entry lost to a truncated save — and those are adopted as orphans.
#
# HUNDREDS means the WRONG STORE IS LOADED, and adopting them is precisely how a stale
# store destroys a healthy board. read_back_human returns ONLY human fields (status,
# applied_date, notes, priority_override), so an adopted row carries a status and NO
# company, role or URL. It renders as a BLANK ROW, and it is saved back into the store,
# so it returns on every later run.
#
# That is exactly what happened on 2026-09-05: a bare `python curate.py` on the VPS
# loaded the stale vault copy (767 entries) instead of the live store (1,471), adopted
# ~700 Sheet rows as identity-less ghosts and blanked the board — including the Shopify
# Offer row and three submitted Tesla applications, which kept their status and lost
# their names. 496 of that store's 767 entries ended up being ghosts.
#
# So: adopt a few, ABORT on many. Aborting costs one refresh. Adopting costs the board.
ORPHAN_ADOPT_MAX = 25

# "Applied 2026-08-29. Shared Tesla resume as-is ..." — the date he typed into the note
# when the Applied column was not there to hold it. Recovered by step 1c.
_APPLIED_IN_NOTE = re.compile(r"\bApplied\s+(\d{4}-\d{2}-\d{2})", re.IGNORECASE)


class StoreMismatch(RuntimeError):
    """The loaded store disagrees with the rendered board badly enough that continuing
    would destroy data. Raised BEFORE anything is written."""


def _merge_readback(store, back: dict, source: str) -> int:
    """Merge a human-field read-back into the store. Returns the orphans adopted."""
    unknown = {cid: h for cid, h in back.items() if cid not in store.postings}

    if len(unknown) > ORPHAN_ADOPT_MAX:
        raise StoreMismatch(
            f"{len(unknown)} of {len(back)} rows on the {source} are absent from the "
            f"loaded store.\n"
            f"    A board is rendered FROM the store and postings are never deleted, so "
            f"this cannot happen with the right file.\n"
            f"    store : {getattr(store, 'path', STORE_PATH)} "
            f"({len(store.postings)} postings)\n"
            f"    {source:6}: {len(back)} rows read back\n"
            f"    NOTHING WAS WRITTEN. The live store is "
            f"~/.hermes/internship/curated_postings.json on the VPS — run the refresh "
            f"via ~/.hermes/scripts/run_curate_vps.sh, or set CURATED_STORE explicitly.")

    for cid, human in back.items():
        if cid in store.postings:
            prev = ((store.postings[cid].get("human") or {}).get("status") or "").strip()
            if prev and prev.lower() != "to apply" and human.get("status", prev) == "":
                # The Sheet is authoritative, blank included — but a cleared Status on a
                # row that HAD an application status is more likely a fat finger on the
                # phone than a decision, so say it out loud. The row returns to Apply Now;
                # `board.py status <m> "<Status>"` puts it back.
                print(f"[refresh] ⚠️  Status CLEARED on the {source} for {cid[:60]} "
                      f"(was {prev!r}) — row returns to the queue", file=sys.stderr)
            store.set_human(cid, human)
    for cid, human in unknown.items():
        store.add_orphan(cid, human)
        print(f"[refresh] ⚠️  orphan adopted from {source}: {cid[:64]} "
              f"(status={human.get('status', '')!r}) — no company/role, so it is kept "
              f"in the store but NOT rendered", file=sys.stderr)
    return len(unknown)


def _touched(h: dict) -> bool:
    """He has done something with this row — it must survive any dedup."""
    st = (h.get("status") or "").strip().lower()
    return bool(st and st != "to apply") or bool((h.get("notes") or "").strip()) \
        or bool((h.get("priority_override") or "").strip())


def _dup_quality(e: dict) -> tuple:
    """Higher sorts first. Employer-hosted beats aggregator, then fuller JD, then fresher."""
    m = e.get("machine") or {}
    url = (m.get("url") or "").lower()
    aggregator = any(a in url for a in ("simplify.jobs", "jobright", "swelist", "api.smartrecruiters"))
    brand = (m.get("source") == "brand-board")
    age = m.get("age_days")
    return (int(brand), int(not aggregator), len(m.get("full_jd") or ""),
            -(999 if age is None else int(age)))


def _norm_rid(m: dict) -> str:
    """Requisition id, comparable across portals. Workday appends -1/-2 to the SAME req on
    its /search/ portal (CIBC …_2617782 vs …_2617782-1); Amazon aggregators often carry no
    req_id at all but the job number is right there in the URL (amazon.jobs/…/jobs/10529525)."""
    rid = str(m.get("req_id") or "").strip().lower()
    if not rid:
        u = (m.get("url") or "").lower()
        hit = re.search(r"amazon\.jobs/(?:[a-z]{2}/)?jobs/(\d{6,})", u)
        if not hit and "myworkdayjobs.com" in u:
            # Workday paths end in _<req id>, with -1/-2 appended on the /search/ portal.
            # Aggregator rows arrive without req_id but with exactly this URL.
            hit = re.search(r"_([a-z]*-?\d[a-z0-9-]*)/?$", u)
        rid = hit.group(1) if hit else ""
    # Strip ONLY a short Workday portal suffix (-1, -2 …) and ONLY when a real id remains.
    # The first version stripped any -\d+ tail and turned Campbell's "Req-66015",
    # "Req-65842", … into a bare "Req" — four DIFFERENT requisitions grouped as one. Caught
    # by the dry run on the live store before it ever ran for real.
    return re.sub(r"^(.*\d.*?)-\d{1,2}$", r"\1", rid)


def _same_human(store, cids) -> bool:
    """True when every row carries IDENTICAL human data, so collapsing loses nothing. This
    is how one req annotated four times with the same On-Hold note collapses to one row."""
    seen = set()
    for c in cids:
        h = store.postings[c].get("human") or {}
        seen.add(tuple(sorted((k, (v or "").strip()) for k, v in h.items() if (v or "").strip())))
    return len(seen) == 1


def _collapse_duplicates(store) -> int:
    """Mark twins of one requisition dead, keeping the best one. Returns rows newly marked.
    Two passes share one loop: (1) same company + same req id — survives the ATS re-titling
    or re-locating a role per portal; (2) same company + title + location for URL variants
    that carry no req id. Idempotent; a revived twin is re-collapsed next run."""
    groups: dict[tuple, list[str]] = {}
    for cid, e in store.postings.items():
        m = e.get("machine") or {}
        if m.get("dead") or not m.get("company") or not m.get("role"):
            continue
        # A row he has already Skipped / Not-a-Fit'd / Closed is resolved: it lives on
        # Reviewed and is not a duplicate candidate (otherwise a hand-skipped twin would
        # trip the "different notes" warning on every run forever).
        if ((e.get("human") or {}).get("status") or "").strip().lower() in REVIEWED_STATUSES:
            continue
        co = normalize_company_name(m["company"])
        rid = _norm_rid(m)
        if rid:
            groups.setdefault(("rid", co, rid), []).append(cid)
        groups.setdefault(("title", co, re.sub(r"\s+", " ", m["role"].strip().lower()),
                           re.sub(r"\s+", " ", (m.get("location") or "").strip().lower())), []).append(cid)

    collapsed = 0
    for key, cids in groups.items():
        cids = [c for c in cids if not store.postings[c]["machine"].get("dead")]
        if len(cids) < 2:
            continue
        if key[0] == "title":
            # Two DIFFERENT known req ids under one title+location are two requisitions
            # (AMD 90891 vs 90947), not one posting seen twice.
            if len({_norm_rid(store.postings[c]["machine"]) for c in cids} - {""}) > 1:
                continue
        touched = [c for c in cids if _touched(store.postings[c].get("human") or {})]
        if len(touched) > 1 and not _same_human(store, touched):
            print(f"[refresh] ⚠️  duplicate rows carry DIFFERENT notes/statuses, leaving all: "
                  f"{key[1]} / {str(key[2])[:50]} — resolve by hand with board.py", file=sys.stderr)
            continue
        pool = touched or cids
        keep = max(pool, key=lambda c: _dup_quality(store.postings[c]))
        for c in cids:
            if c == keep:
                continue
            mm = store.postings[c]["machine"]
            mm["dead"] = True
            mm["dead_reason"] = f"duplicate of {keep}"
            collapsed += 1
    return collapsed


async def refresh(notify: bool = False) -> int:
    # An open-in-Excel lock no longer aborts the whole refresh — we still harvest, score,
    # and update the store JSON, and only DEFER the xlsx render (step 5). That keeps the
    # board data fresh even when it's left open in Excel for days (the old behavior froze
    # the board until Excel was closed). read_back_human can still read an open .xlsx.
    _wait_for_network()      # scheduled runs fire on wake before wifi is up — wait for it
    today = date.today().isoformat()
    store = CuratedStore(STORE_PATH).load()
    prior_cids = set(store.postings.keys())   # to detect genuinely-new postings
    # Pre-merge routing snapshot for step 1c. A date may be stamped ONLY on a row that
    # BECOMES an application during this run; a row that was already an undated
    # application before the run has an unknown real date, and today is not it.
    _was_application = {
        cid: classify_row({"_cid": cid, "machine": e.get("machine") or {},
                           "human": e.get("human") or {}}) == "application"
        for cid, e in store.postings.items()}

    # 1) read his edits back from the Google Sheet — THE source of truth for human fields
    # (Sparsh, 2026-09-05: "we no longer use any mac xlsx sheet, only the google sheet").
    # The xlsx read-back that used to run first is GONE. On the VPS the xlsx is rendered
    # from this same store one run earlier, so reading it back could only re-inject a
    # value he had since cleared on the Sheet (both read-backs skipped blank cells, so a
    # cleared cell could never stick). The xlsx is now write-only output.
    if gsheet is not None and GSHEET_ON:
        try:
            g_back = gsheet.read_back_human(GSHEET_ID)
            _merge_readback(store, g_back, "Sheet")
            if g_back:
                print(f"[refresh] merged Sheet edits on {len(g_back)} rows", file=sys.stderr)
        except Exception as e:  # noqa: BLE001
            print(f"⚠️  [refresh] could not read the Google Sheet: "
                  f"{type(e).__name__}: {e}", file=sys.stderr)

    # 1c) BACKFILL A MISSING APPLICATION DATE (added 2026-09-05).
    # board.py writes the date into the Sheet, but the Apply Now tab HAS NO "Applied"
    # COLUMN — and _set() returns silently when a header is absent. So marking a row
    # Applied while it was still in the queue could never record a date, by either
    # command, and nothing said so. 15 real applications (9 Tesla, 6 Microsoft) were
    # sitting undated when this was found.
    # Stamp the date the first time a refresh sees an application status with no date.
    # This is the REFRESH date, not necessarily the submit date (it can be up to one
    # refresh interval late), so it only ever FILLS A BLANK and never overwrites a
    # real value. Routing comes from classify_row() so this cannot drift from the
    # definition of "is an application" the board itself uses.
    # 🔴 Two rules, learned the same evening this was added. (1) A date written in the
    # Notes ("Applied 2026-08-29 ...") is EVIDENCE — use it first. (2) Otherwise stamp
    # the refresh date ONLY on a row that transitioned into an application in THIS run.
    # The first version stamped every undated application with today's date and
    # back-dated rejections from August to September 5th: an unmarked guess written as
    # fact, which Fact Hygiene rule 1 forbids. An old undated row stays blank.
    dated = recovered = 0
    for _cid, _e in store.postings.items():
        _h = _e.get("human") or {}
        if (_h.get("applied_date") or "").strip():
            continue
        if classify_row({"_cid": _cid, "machine": _e.get("machine") or {},
                         "human": _h}) != "application":
            continue
        _hit = _APPLIED_IN_NOTE.search(_h.get("notes") or "")
        if _hit:
            store.set_human(_cid, {"applied_date": _hit.group(1)})
            recovered += 1
        elif not _was_application.get(_cid, False):
            store.set_human(_cid, {"applied_date": today})
            dated += 1
    if dated or recovered:
        print(f"[refresh] applied_date: {dated} stamped (newly applied this run), "
              f"{recovered} recovered from Notes (see 1c — the queue tab cannot carry "
              f"the column)", file=sys.stderr)

    # 2) one-time history seed
    seeded = import_old_tracker_history(store)
    if seeded:
        print(f"[refresh] seeded {seeded} history rows from old tracker", file=sys.stderr)

    # 3) harvest — lane 1 (brand-first boards) + lane 2 (wide net: aggregators + Gmail)
    print("[refresh] lane 1: pulling target-company boards…", file=sys.stderr)
    lane1 = await brand_first_source.collect()
    for r in lane1:
        r["_src"] = "brand-board"
    print("[refresh] lane 2: wide net (aggregators + Gmail)…", file=sys.stderr)
    lane2 = await wide_net_source.collect()
    for r in lane2:
        r["_src"] = r["source"]

    # clean display company names (strip aggregator emoji/symbols) + cross-lane dedup
    # by (company, role): same role from two sources -> keep the better one
    # (brand-board > more JD > fresher).
    for r in lane1 + lane2:
        r["company"] = _clean_company(r["company"])

    def _better(a, b):
        a1, b1 = a["_src"] == "brand-board", b["_src"] == "brand-board"
        if a1 != b1:
            return a1
        if len(a.get("full_jd", "")) != len(b.get("full_jd", "")):
            return len(a.get("full_jd", "")) > len(b.get("full_jd", ""))
        # `or 999` is WRONG here: age_days == 0 is falsy, so a role posted TODAY
        # scored 999 and LOST dedup to an older duplicate — exactly inverted, and
        # it penalised precisely the postings most worth applying to first.
        a_age = a.get("age_days")
        b_age = b.get("age_days")
        return (999 if a_age is None else a_age) < (999 if b_age is None else b_age)

    best: dict[tuple, dict] = {}
    for r in lane1 + lane2:
        key = (normalize_company_name(r["company"]), r["role"].strip().lower())
        if key not in best or _better(r, best[key]):
            best[key] = r

    harvested: set[str] = set()
    for r in best.values():
        cid = r["canonical_id"]
        harvested.add(cid)
        _score_into(store, cid, {
            "company": r["company"], "role": r["role"], "location": r["location"],
            "url": r["url"], "ats_type": r["ats_type"], "cycle": r["cycle"],
            "posted_date": r["posted_date"], "age_days": r["age_days"],
            "full_jd": r["full_jd"], "req_id": r.get("req_id", ""),
        }, first_seen=today, source=r["_src"])
    print(f"[refresh] {len(lane1)} lane-1 + {len(lane2)} lane-2 -> "
          f"{len(harvested)} after cross-lane dedup", file=sys.stderr)

    # SAFETY GUARD: a network-less cron run (DNS failures on a sleeping/just-woke Mac)
    # harvests ~nothing. WITHOUT this, the stale-check below would mark every posting
    # dead and wipe the board to 0. If the harvest collapsed but we had a healthy board
    # before, abort and leave the store + xlsx UNTOUCHED (the next good run restores it).
    # A healthy run harvests 100+ roles. Under 50 (when we had a real board before) means
    # a network failure (full or partial), not that the postings genuinely vanished.
    if len(harvested) < 50 and len(prior_cids) > 80:
        print(f"[refresh] ⚠️ harvest collapsed ({len(harvested)} roles vs {len(prior_cids)} "
              f"stored) — almost certainly a network failure. Aborting WITHOUT touching the "
              f"board (no stale-check, no overwrite).", file=sys.stderr)
        return 0

    # 4) re-score everything (recency decays) + stale-check dropouts
    #
    # 2026-08-18 REWRITE — `dead` used to mean "my source stopped listing it",
    # which is NOT the same as "the job closed". Two independent false-dead
    # mechanisms were measured against employers' own ATS APIs:
    #   • wide-net roll-off: the GitHub/Gmail aggregators are ROLLING WINDOWS and
    #     drop older entries to stay readable. 54% of dead wide-net rows were
    #     still open (Palantir x3, IMC Summer 2027, Modal, Binance.US...).
    #   • brand-board fetch failure: a transient timeout looked identical to an
    #     empty board. 14% of dead brand-board rows were still open, incl.
    #     Anduril "2027 Software Engineer Intern".
    # Fixes, in order of how much they recover:
    #   (a) postings from a board that ERRORED this run are exempt entirely
    #   (b) `manual` boards (Google/Meta/Microsoft/Tesla/Apple/Netflix/Uber/
    #       Rippling) are never auto-dead — lane 1 can't fetch them at all, so a
    #       disappearance carries no information. They're the top brands.
    #   (c) wide-net rows need far more strikes, because roll-off is expected
    #   (d) PER-LANE collapse guard: the old guard only tested the COMBINED
    #       harvest (<50), so when lane 1 collapsed and lane 2 stayed healthy the
    #       total cleared the bar and every brand-board posting took a strike.
    #       That happened in 14 of 108 logged runs (13%) — including lane1=5
    #       against a healthy median of 68.
    failed_boards = {normalize_company_name(n)
                     for n in getattr(brand_first_source, "FAILED_BOARDS", set())}
    manual_boards = {normalize_company_name(b["name"]) for b in _boards()
                     if b.get("ats_type") == "manual"}

    lane1_ok = len(lane1) >= LANE1_MIN_HEALTHY
    if not lane1_ok:
        print(f"[refresh] ⚠️ lane-1 harvest is anomalously low ({len(lane1)} < "
              f"{LANE1_MIN_HEALTHY}) — brand-board postings are EXEMPT from the "
              f"stale-check this run (partial-collapse guard).", file=sys.stderr)

    stale = 0
    for cid, rec in store.items():
        m = rec.get("machine", {})
        src = m.get("source", "")
        is_brand = src == "brand-board"
        is_wide = src.startswith("wide:")
        nn = normalize_company_name(m.get("company", ""))

        exempt = (
            nn in failed_boards            # (a) its board errored this run
            or nn in manual_boards         # (b) no API exists to confirm death
            or (is_brand and not lane1_ok)  # (d) lane-1 partial collapse
        )
        strikes_needed = WIDE_STALE_STRIKES if is_wide else STALE_STRIKES

        if (is_brand or is_wide) and cid not in harvested and not exempt:
            m["fail_count"] = int(m.get("fail_count", 0)) + 1
            if m["fail_count"] >= strikes_needed and not m.get("dead"):
                m["dead"] = True
                stale += 1
        elif cid in harvested:
            m["fail_count"] = 0
            m["dead"] = False               # reappeared -> it was never dead
        # refresh age + re-score from stored posted_date
        if m.get("company") and m.get("role"):
            age = brand_first_source._age_from_date(m.get("posted_date", ""))
            h = hotness(m["company"], m["role"], age, lane=role_lane(m["role"]))
            m.update({"age_days": age, "hotness": h["hotness"], "fresh": h["fresh"],
                      "recency": h["recency"], "tier": h["tier"], "brand": h["brand"],
                      "lane": h["lane"], "role_score": h["role"]})

    # 4a) COLLAPSE URL-VARIANT DUPLICATES (added 2026-09-05).
    # canonical_id() keys on the URL, and ATSes hand out several URLs for one req:
    # Workday with and without `/en-US/`, path case (`/DMA/` vs `/dma/`), an
    # aggregator link (simplify.jobs/p/…) beside the employer's own, SmartRecruiters'
    # api.* beside jobs.*. Each variant minted its own row: 19 live (company, role)
    # pairs were on the board twice. Changing canonical_id would re-key every stored
    # posting (a migration), so the fix is here: the twins stay in the store and one
    # is marked dead with `dead_reason` naming its keeper — no id changes, no data loss.
    # 🔴 A row he has TOUCHED (any status beyond To Apply, a note, a priority) is never
    # the one that dies; if both twins are touched, both stay and it is logged.
    collapsed = _collapse_duplicates(store)
    if collapsed:
        print(f"[refresh] collapsed {collapsed} URL-variant duplicate rows "
              f"(kept the touched/brand/fuller twin; dead_reason names it)", file=sys.stderr)

    # 4b) FIT PASS — the AI reads each JD: fit score + why + disqualifier flag.
    #     Cached by JD-hash, so only new/changed postings cost anything (idempotent).
    try:
        from fit_pass import run_fit_pass
        fs = run_fit_pass(store)
        print(f"[refresh] fit: {fs.scored} scored, {fs.cached} cached, "
              f"{fs.no_jd} no-JD, {fs.errors} errors · ~${fs.cost:.4f}", file=sys.stderr)
    except Exception as e:  # noqa: BLE001 — never block the refresh on the LLM
        print(f"[refresh] fit pass skipped: {type(e).__name__}: {e}", file=sys.stderr)

    # 5) persist the store ALWAYS; render the xlsx unless Excel has it open (defer if so)
    gen = datetime.now().strftime("%Y-%m-%d %H:%M")
    store.save(gen)
    if is_locked(XLSX_PATH):
        print(f"\n⏸️  Store updated ({len(store.postings)} postings, {stale} newly stale) — xlsx "
              f"render DEFERRED ('{XLSX_PATH.name}' is open in Excel). The board re-renders on the "
              f"next run with Excel closed; no data lost, notify still runs.", file=sys.stderr)
    else:
        counts = write_board(store, XLSX_PATH, gen)
        print(f"\n✅ Curated Board refreshed — {counts['queue']} in queue, "
              f"{counts['applications']} applications, {counts.get('reviewed', 0)} reviewed/skipped "
              f"({len(harvested)} live brand-board roles, {stale} newly stale)")
        print(f"   {XLSX_PATH}")

    # 5b) the Google Sheet. Isolated on purpose: a network blip, an expired OAuth or a
    # Sheets outage must never cost us the store save or the xlsx above. It fails LOUD
    # (a visible warning) rather than silently, because a board that has quietly stopped
    # updating looks exactly like a board with nothing new on it.
    if gsheet is not None and GSHEET_ON:
        try:
            res = gsheet.write_board(store.postings, GSHEET_ID, gen)
            late = res.get("human") or {}
            changed = 0
            for cid, human in late.items():          # edits typed during the harvest
                if cid not in store.postings:
                    continue
                # Compare FIELD BY FIELD. read_back_human returns only the non-empty
                # cells, so comparing whole dicts against the store's full human record
                # reports every row as changed on every run.
                cur = (store.get(cid) or {}).get("human") or {}
                if any((cur.get(k) or "") != v for k, v in human.items()):
                    store.set_human(cid, human)
                    changed += 1
            if changed:
                store.save(gen)
            c = res["counts"]
            print(f"   Google Sheet updated — {c[gsheet.TAB_QUEUE]} queue, "
                  f"{c[gsheet.TAB_APPS]} applications, {c[gsheet.TAB_REVIEWED]} reviewed"
                  + (f" (+{changed} late edits merged)" if changed else ""))
        except Exception as e:  # noqa: BLE001
            print(f"\n⚠️  GOOGLE SHEET NOT UPDATED — {type(e).__name__}: {e}\n"
                  f"    The xlsx and the store are fine. The Sheet is now STALE.",
                  file=sys.stderr)

    # 6) notify (only when scheduled, only on genuinely-new postings, never on the
    #    first-ever run; silent otherwise — matches the other crons' discipline)
    new_cids = harvested - prior_cids
    new_live = {c for c in new_cids
                if not (store.get(c) or {}).get("machine", {}).get("dead")}
    if notify and prior_cids and new_live:
        msg = _new_postings_digest(store, new_live)
        if msg and _send_telegram(msg):
            print(f"[notify] sent Telegram digest: {len(new_live)} new role(s)", file=sys.stderr)
    elif notify:
        print(f"[notify] {len(new_live)} new (prior_store={bool(prior_cids)}) — "
              f"no message sent", file=sys.stderr)
    return 0


async def validate_boards() -> int:
    import ats_router as A
    from company_boards import boards
    print(f"{'COMPANY':22} {'ATS':16} {'ROLES':>6} {'INTERN':>7}")
    dead = []
    async with A.make_client() as c:
        async def chk(b):
            if b["ats_type"] == "manual":
                return b["name"], "manual", 0, 0
            try:
                recs = await A.fetch_board(c, b)
                return (b["name"], b["ats_type"], len(recs),
                        len([r for r in recs if A.default_intern_filter(r.title)]))
            except Exception as e:  # noqa: BLE001
                return b["name"], f"ERR {type(e).__name__}", -1, -1
        for name, ats, roles, interns in await asyncio.gather(*[chk(b) for b in boards()]):
            flag = "✓" if roles > 0 else ("· manual" if ats == "manual" else "✗ DEAD")
            if roles == 0 and ats != "manual":
                dead.append(name)
            print(f"{name:22} {ats:16} {roles:>6} {interns:>7}  {flag}")
    if dead:
        print("\nDEAD/EMPTY — fix token in company_boards.py:", ", ".join(dead))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Refresh the curated brand-first board.")
    ap.add_argument("--validate-boards", action="store_true",
                    help="Ping every board and report role counts (no write).")
    ap.add_argument("--notify", action="store_true",
                    help="Send a Telegram digest if genuinely-new postings appeared "
                         "(silent otherwise). Used by the scheduled run.")
    args = ap.parse_args()
    if args.validate_boards:
        return asyncio.run(validate_boards())
    try:
        return asyncio.run(refresh(notify=args.notify))
    except StoreMismatch as e:
        # Exit 2, not a traceback: this is an operator error with a known remedy, and
        # the whole point of the guard is that the message gets read.
        print(f"\n🛑 REFRESH ABORTED — store/board mismatch.\n   {e}\n", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())

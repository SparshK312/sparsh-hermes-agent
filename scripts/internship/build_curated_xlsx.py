#!/usr/bin/env python3
"""
build_curated_xlsx.py — the living two-sheet board (round-trip preserve).

THE CRUX: regenerate the xlsx fresh every refresh (so ranking re-sorts and all
formatting is re-emitted from code = nothing can be lost), while the user's manual
edits (Status / Applied / Notes / Priority) survive — because they're read back out
of the OLD xlsx keyed by a hidden canonical_id column and merged into the JSON store
BEFORE the new file is written.

  read_back_human(path)  -> {canonical_id: {status, applied_date, notes, priority_override}}
  check_lock(path)       -> raise if Excel has the file open (~$ lock)
  write_board(store, path, generated_at)

Two sheets, driven from curated_store:
  🔥 Curated Queue   — machine-ranked discovery; Status in {"", "To Apply"}
  📋 My Applications — anything the user has actioned (Status set to else)
Plus a Summary (COUNTIFs) and a hidden _meta sheet (schema fingerprint).

openpyxl note: we NEVER save a loaded workbook. read_back opens read-only and only
reads; write_board builds a brand-new Workbook(). So data-validation / conditional-
formatting / hyperlinks are guaranteed by construction, not preserved-through-load.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

SCHEMA_VERSION = 3
ID_HEADER = "_id"

# ── dropdown vocab ────────────────────────────────────────────────────────────
STATUS_OPTS = ["To Apply", "Applied", "OA", "Phone Screen", "Onsite", "Offer",
               "Rejected", "Networking", "On Hold", "Skip", "Not a Fit", "Closed"]
# Statuses that mean "I looked at this and I'm not applying" -> Reviewed sheet, not the
# active queue and NOT the applications sheet. Skip / Not a Fit = your judgment call;
# Closed = dead (missed deadline / role pulled) that you never applied to.
REVIEWED_STATUSES = {"skip", "not a fit", "closed"}
PRIORITY_OPTS = ["", "P0", "P1", "P2", "P3"]
LANE_OPTS = ["AI/ML", "SWE", "Data", "PM", "Other"]
CYCLE_OPTS = ["Fall 2026", "Winter 2027", "Spring 2027", "Summer 2027", "Summer 2026", "TBD"]

# ── palette (matches the prior tracker) ───────────────────────────────────────
SLATE = "1F2937"; WHITE = "FFFFFF"
# Distinct hue per status so the Status column reads at a glance. Progression of an
# active application moves warm->cool->green (To Apply -> ... -> Offer); parked/dead
# states are muted. bg = ~Tailwind-200, text = ~800 (readable in Excel).
STATUS_FILL = {"To Apply": ("FED7AA", "9A3412"),        # orange  — action needed
               "On Hold": ("E2E8F0", "334155"),         # slate   — parked / waiting on something
               "Applied": ("BFDBFE", "1E40AF"),         # blue    — in the pipeline
               "OA": ("BAE6FD", "075985"),              # sky
               "Phone Screen": ("C7D2FE", "3730A3"),    # indigo
               "Onsite": ("DDD6FE", "5B21B6"),          # violet
               "Offer": ("A7F3D0", "065F46"),           # green   — win
               "Networking": ("FBCFE8", "9D174D"),      # pink    — warm outreach
               "Rejected": ("FECACA", "991B1B"),        # red     — no
               "Skip": ("F3F4F6", "6B7280"),            # cool gray
               "Not a Fit": ("E7E5E4", "57534E"),       # warm gray
               "Closed": ("FEE2E2", "B91C1C")}          # light red — dead/deadline
TIER_FILL = {"S": ("FEF3C7", "92400E"), "A": ("D1FAE5", "065F46"),
             "B": ("DBEAFE", "1E3A8A"), "C": ("F3F4F6", "6B7280")}
PRIO_FILL = {"P0": "FDE68A", "P1": "BBF7D0", "P2": "BFDBFE", "P3": "E5E7EB"}

_thin = Side(style="thin", color="E5E7EB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

PRIO_RANK = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "": 4}
STATUS_RANK = {"Offer": 0, "Onsite": 1, "Phone Screen": 2, "OA": 3,
               "Applied": 4, "Networking": 5, "On Hold": 6, "Rejected": 9}


def classify_row(rec: dict) -> str:
    """Which sheet a posting belongs on: 'queue' | 'application' | 'reviewed' | 'drop'.
    - queue:       live, still 'To Apply' (machine-ranked discovery).
    - application: any real application status (Applied/OA/Onsite/Offer/Rejected/...).
    - reviewed:    Skip/Not-a-Fit you set, OR a posting that went stale while you had a
                   note on it (so your note is never silently lost — that's the whole point).
    - drop:        stale with no note and no status -> off the board (still kept in JSON).
    """
    m, h = rec.get("machine", {}), rec.get("human", {})
    status = (h.get("status") or "").strip().lower()
    note = (h.get("notes") or "").strip()
    if status in REVIEWED_STATUSES:
        return "reviewed"
    # "On Hold" = parked before applying (e.g. waiting on a referral) -> stays in the
    # queue, NOT the applications sheet (nothing has actually been applied to yet).
    if status not in ("", "to apply", "on hold"):
        return "application"
    if m.get("dead"):
        return "reviewed" if note else "drop"
    return "queue"


# ── lock detection ────────────────────────────────────────────────────────────
def is_locked(path: str | Path) -> bool:
    """True if Excel has the file open (its ~$ owner-lock is present). Non-raising —
    callers use this to DEFER the xlsx render (not abort the whole refresh)."""
    p = Path(path)
    return p.with_name("~$" + p.name).exists()


def check_lock(path: str | Path) -> None:
    if is_locked(path):
        p = Path(path)
        raise RuntimeError(
            f"'{p.name}' looks open in Excel (lock file ~${p.name} present). "
            f"Close the spreadsheet and re-run.")


# ── read-back ─────────────────────────────────────────────────────────────────
_HUMAN_BY_HEADER = {"status": "status", "priority": "priority_override",
                    "applied": "applied_date", "notes": "notes"}


def read_back_human(path: str | Path) -> dict[str, dict]:
    """Pull human-edited fields out of the existing xlsx, keyed by canonical_id
    (hidden _id column). Reads both data sheets. Never saves."""
    p = Path(path)
    if not p.exists():
        return {}
    out: dict[str, dict] = {}
    wb = load_workbook(p, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            if ws.title.startswith("_") or ws.title == "Summary":
                continue
            header_row = None
            col_of: dict[str, int] = {}
            for row in ws.iter_rows(values_only=True):
                if header_row is None:
                    if row and ID_HEADER in [str(c).strip() if c else "" for c in row]:
                        header_row = [str(c).strip() if c else "" for c in row]
                        for idx, h in enumerate(header_row):
                            col_of[h.lower()] = idx
                    continue
                idx_id = col_of.get(ID_HEADER.lower())
                if idx_id is None or idx_id >= len(row):
                    continue
                cid = row[idx_id]
                if not cid:
                    continue
                human = out.setdefault(str(cid), {})
                for hname, field in _HUMAN_BY_HEADER.items():
                    ci = col_of.get(hname)
                    if ci is not None and ci < len(row) and row[ci] not in (None, ""):
                        human[field] = str(row[ci]).strip()
    finally:
        wb.close()
    return out


# ── styling helpers ───────────────────────────────────────────────────────────
def _title_block(ws, ncols, title, subtitle):
    last = get_column_letter(ncols)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{last}1")
    ws["A1"].value = title
    ws["A1"].font = Font(size=15, bold=True, color="111827")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A2:{last}2")
    ws["A2"].value = subtitle
    ws["A2"].font = Font(size=9, italic=True, color="6B7280")
    ws.row_dimensions[2].height = 15


def _header(ws, headers, widths, hdr_row):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.font = Font(bold=True, color=WHITE, size=10.5)
        c.fill = PatternFill("solid", fgColor=SLATE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
    ws.row_dimensions[hdr_row].height = 22
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.column_dimensions["A"].hidden = True   # _id column


def _add_dv(ws, col, options, first, last):
    dv = DataValidation(type="list", formula1='"' + ",".join(o for o in options if o) + '"',
                        allow_blank=True)
    dv.add(f"{col}{first}:{col}{last}")
    ws.add_data_validation(dv)


# ── sheet builders ────────────────────────────────────────────────────────────
# Fit + Why inserted after Hot. Column positions are looked up via COL (below) so
# adding/moving a column never requires hunting magic indices.
QUEUE_HEADERS = [ID_HEADER, "🔥", "Hot", "Fit", "Why", "Tier", "Priority", "Status",
                 "Company", "Role", "Lane", "Location", "Cycle", "Posted", "Age",
                 "Apply", "Source", "Notes"]
QUEUE_WIDTHS = [2, 4, 6, 6, 32, 6, 9, 13, 16, 40, 8, 18, 12, 11, 6, 8, 12, 28]
QCOL = {h: i + 1 for i, h in enumerate(QUEUE_HEADERS)}   # header -> 1-based column

DISQ_FILL = ("F3F4F6", "B91C1C")   # grey bg, red font for disqualified Fit cell
NOJD_FILL = ("FEF3C7", "92400E")   # amber — "AI couldn't read this, check it yourself"


def _is_disq(m: dict) -> bool:
    return (m.get("fit_disqualifier") or "none") not in ("none", "", None)


def _fit_cells(m: dict):
    """Return (fit_display, why_display, kind) for a machine record.
    kind ∈ {'disq','scored','nojd','none'} drives the cell styling."""
    dq = m.get("fit_disqualifier") or "none"
    score = m.get("fit_score")
    why = m.get("fit_why") or ""
    if dq not in ("none", "", None):
        return "❌", f"❌ {dq} — {why}".strip(" —"), "disq"
    if isinstance(score, (int, float)) or (isinstance(score, str) and str(score).isdigit()):
        return int(score), why, "scored"
    if not (m.get("full_jd") or "").strip():
        return "👀", "⚠️ AI couldn't read this — no JD (bot-blocked). Click Apply to check it yourself.", "nojd"
    return "", "", "none"


def _queue_sort_key(rec):
    m, h = rec["machine"], rec["human"]
    prio = (h.get("priority_override") or "").strip()
    on_hold = 1 if (h.get("status") or "").strip().lower() == "on hold" else 0
    # within a priority band: active To-Apply first, then On Hold (parked), then
    # disqualified rows sink to the bottom -- all still visible/filterable.
    return (PRIO_RANK.get(prio, 4), 1 if _is_disq(m) else 0, on_hold,
            -int(m.get("hotness", 0) or 0))


def _build_queue(ws, rows):
    _title_block(ws, len(QUEUE_HEADERS),
                 "🔥 Curated Queue — brand-ranked open intern roles",
                 "Ranked by Hotness (brand-dominant). Fit = AI read of the JD (hover a Role "
                 "cell for the JD). ❌ = disqualified (still shown, sunk). Set Status to move a row.")
    hdr = 3; first = 4
    _header(ws, QUEUE_HEADERS, QUEUE_WIDTHS, hdr)
    rows = sorted(rows, key=_queue_sort_key)
    L = get_column_letter
    CENTER = {QCOL["🔥"], QCOL["Hot"], QCOL["Fit"], QCOL["Tier"], QCOL["Age"]}
    WRAP = {QCOL["Why"], QCOL["Role"], QCOL["Notes"]}
    for i, rec in enumerate(rows):
        r = first + i
        m, h = rec["machine"], rec["human"]
        fit_disp, why_disp, fit_kind = _fit_cells(m)
        vals = {
            ID_HEADER: rec["_cid"], "🔥": m.get("fresh", ""), "Hot": m.get("hotness", ""),
            "Fit": fit_disp, "Why": why_disp, "Tier": m.get("tier", ""),
            "Priority": h.get("priority_override", ""), "Status": h.get("status", "") or "To Apply",
            "Company": m.get("company", ""), "Role": m.get("role", ""), "Lane": m.get("lane", ""),
            "Location": m.get("location", ""), "Cycle": m.get("cycle", ""),
            "Posted": m.get("posted_date", ""), "Age": m.get("age_days", ""),
            "Apply": ("Apply ↗" if m.get("url") else ""), "Source": m.get("source", ""),
            "Notes": h.get("notes", ""),
        }
        for hname, col in QCOL.items():
            c = ws.cell(row=r, column=col, value=vals[hname])
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(col in WRAP),
                                    horizontal=("center" if col in CENTER else "left"))
        if i % 2 == 1:
            for j in range(1, len(QUEUE_HEADERS) + 1):
                if ws.cell(row=r, column=j).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="F9FAFB")
        # tier fill
        tf = TIER_FILL.get(m.get("tier", ""))
        if tf:
            tc = ws.cell(row=r, column=QCOL["Tier"])
            tc.fill = PatternFill("solid", fgColor=tf[0]); tc.font = Font(bold=True, size=10, color=tf[1])
        # status fill
        sf = STATUS_FILL.get(vals["Status"])
        if sf:
            sc = ws.cell(row=r, column=QCOL["Status"])
            sc.fill = PatternFill("solid", fgColor=sf[0]); sc.font = Font(bold=True, size=10, color=sf[1])
        # priority fill
        pf = PRIO_FILL.get(vals["Priority"])
        if pf:
            ws.cell(row=r, column=QCOL["Priority"]).fill = PatternFill("solid", fgColor=pf)
        # fit cell styling: disqualified = red ❌; no-JD = faint marker (gradient CF handles scored)
        fc = ws.cell(row=r, column=QCOL["Fit"])
        if fit_kind == "disq":
            fc.fill = PatternFill("solid", fgColor=DISQ_FILL[0]); fc.font = Font(bold=True, size=11, color=DISQ_FILL[1])
        elif fit_kind == "nojd":
            fc.fill = PatternFill("solid", fgColor=NOJD_FILL[0])
            fc.font = Font(size=11, bold=True, color=NOJD_FILL[1])
        # JD surfaced as a hover-comment on the Role cell (uses the otherwise-wasted JD)
        jd = (m.get("full_jd") or "").strip()
        if jd:
            summ = (m.get("fit_jd_summary") or "").strip()
            txt = (f"{summ}\n\n— JD —\n{jd[:800]}" if summ else jd[:900])
            cm = Comment(txt, "fit-pass"); cm.width = 460; cm.height = 300
            ws.cell(row=r, column=QCOL["Role"]).comment = cm
        # apply hyperlink
        if m.get("url"):
            ac = ws.cell(row=r, column=QCOL["Apply"])
            ac.hyperlink = m["url"]
            ac.font = Font(size=10, color="2563EB", underline="single")
            ac.alignment = Alignment(horizontal="center", vertical="center")
    last = first + len(rows) - 1 if rows else first
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"B{hdr}:{L(len(QUEUE_HEADERS))}{last}"
    if rows:
        _add_dv(ws, L(QCOL["Priority"]), PRIORITY_OPTS, first, last)
        _add_dv(ws, L(QCOL["Status"]), STATUS_OPTS, first, last)
        _add_dv(ws, L(QCOL["Lane"]), LANE_OPTS, first, last)
        _add_dv(ws, L(QCOL["Cycle"]), CYCLE_OPTS, first, last)
        st_col = L(QCOL["Status"])
        for val, (bg, fg) in STATUS_FILL.items():
            ws.conditional_formatting.add(f"{st_col}{first}:{st_col}{last}",
                CellIsRule(operator="equal", formula=[f'"{val}"'],
                           fill=PatternFill("solid", fgColor=bg),
                           font=Font(color=fg, bold=True, size=10)))
        # Hot + Fit both get the white->green numeric gradient
        for cl in (L(QCOL["Hot"]), L(QCOL["Fit"])):
            ws.conditional_formatting.add(f"{cl}{first}:{cl}{last}",
                ColorScaleRule(start_type="num", start_value=30, start_color="FFFFFF",
                               end_type="num", end_value=100, end_color="34D399"))
    return last


APP_HEADERS = [ID_HEADER, "Status", "Company", "Role", "Lane", "Location", "Cycle",
               "Apply", "Applied", "Source / Referral", "Notes"]
APP_WIDTHS = [2, 13, 18, 40, 8, 20, 12, 8, 12, 18, 38]


def _build_apps(ws, rows):
    _title_block(ws, len(APP_HEADERS),
                 "📋 My Applications — what you've actioned",
                 "Stable tracker (never re-ranked). Rows land here when you set a Status "
                 "other than 'To Apply'. Sorted by stage then applied date.")
    hdr = 3; first = 4
    _header(ws, APP_HEADERS, APP_WIDTHS, hdr)

    def sort_key(rec):
        h = rec["human"]
        return (STATUS_RANK.get((h.get("status") or "").strip(), 7),
                "0" if not h.get("applied_date") else h["applied_date"])
    rows = sorted(rows, key=sort_key)
    for i, rec in enumerate(rows):
        r = first + i
        m, h = rec["machine"], rec["human"]
        vals = [rec["_cid"], h.get("status", ""), m.get("company", ""), m.get("role", ""),
                m.get("lane", ""), m.get("location", ""), m.get("cycle", ""),
                ("Apply ↗" if m.get("url") else ""), h.get("applied_date", ""),
                m.get("source", ""), h.get("notes", "")]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(j in (4, 11)))
        sf = STATUS_FILL.get(vals[1])
        if sf:
            sc = ws.cell(row=r, column=2)
            sc.fill = PatternFill("solid", fgColor=sf[0])
            sc.font = Font(bold=True, size=10, color=sf[1])
        if m.get("url"):
            ac = ws.cell(row=r, column=8)
            ac.hyperlink = m["url"]
            ac.font = Font(size=10, color="2563EB", underline="single")
            ac.alignment = Alignment(horizontal="center", vertical="center")
    last = first + len(rows) - 1 if rows else first
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"B{hdr}:{get_column_letter(len(APP_HEADERS))}{last}"
    if rows:
        _add_dv(ws, "B", STATUS_OPTS, first, last)
        for val, (bg, fg) in STATUS_FILL.items():
            ws.conditional_formatting.add(f"B{first}:B{last}",
                CellIsRule(operator="equal", formula=[f'"{val}"'],
                           fill=PatternFill("solid", fgColor=bg),
                           font=Font(color=fg, bold=True, size=10)))
    return first, last


def _build_summary(wb, app_first, app_last, queue_count, reviewed_count, generated_at):
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["A"].width = 22
    sm.column_dimensions["B"].width = 12
    sm["A1"] = "Curated Board — Summary"
    sm["A1"].font = Font(size=14, bold=True, color="111827")
    sm["A2"] = f"generated {generated_at}"
    sm["A2"].font = Font(size=9, italic=True, color="9CA3AF")
    rng = f"'My Applications'!$B${app_first}:$B${app_last}"
    metrics = [
        ("", ""),
        ("DISCOVERY", None),
        ("Curated queue (to apply)", queue_count),
        ("", ""),
        ("APPLICATIONS", None),
        ("Applied", f'=COUNTIF({rng},"Applied")'),
        ("In process (OA+)", f'=COUNTIF({rng},"OA")+COUNTIF({rng},"Phone Screen")+COUNTIF({rng},"Onsite")'),
        ("Offers", f'=COUNTIF({rng},"Offer")'),
        ("Rejected", f'=COUNTIF({rng},"Rejected")'),
        ("Networking", f'=COUNTIF({rng},"Networking")'),
        ("", ""),
        ("REVIEWED", None),
        ("Skipped / closed", reviewed_count),
    ]
    for i, (label, formula) in enumerate(metrics, start=4):
        sm.cell(row=i, column=1, value=label)
        if formula is None:
            sm.cell(row=i, column=1).font = Font(bold=True, color="6B7280", size=10)
        elif formula != "":
            sm.cell(row=i, column=1).font = Font(size=11)
            vc = sm.cell(row=i, column=2, value=formula)
            vc.font = Font(size=12, bold=True, color="1E3A8A")
            vc.alignment = Alignment(horizontal="center")


REVIEW_HEADERS = [ID_HEADER, "Status", "Company", "Role", "Lane", "Location", "Cycle",
                  "Fit", "Why", "Apply", "Notes"]
REVIEW_WIDTHS = [2, 12, 18, 38, 8, 18, 12, 6, 30, 8, 36]


def _review_status(rec) -> str:
    """Display status on the Reviewed sheet: your Skip/Not-a-Fit if set, else 'Closed'
    for a posting that went stale while you had a note on it. 'To Apply'/blank is not a
    real status, so a dead row carrying it shows as Closed."""
    status = (rec["human"].get("status") or "").strip()
    if status and status.lower() != "to apply":
        return status
    return "Closed" if rec["machine"].get("dead") else status


def _build_reviewed(ws, rows):
    _title_block(ws, len(REVIEW_HEADERS),
                 "🗂️ Reviewed / Skipped — passed on, or closed",
                 "Rows you marked Skip / Not a Fit, plus postings that closed while you "
                 "had notes on them. Your reason notes are kept here, never lost. To send "
                 "one back to the queue, set its Status to 'To Apply' (only works while it's live).")
    hdr = 3; first = 4
    _header(ws, REVIEW_HEADERS, REVIEW_WIDTHS, hdr)

    def sort_key(rec):
        # Skips first, then Closed (stale); alpha by company within each.
        ds = _review_status(rec)
        return (0 if ds in ("Skip", "Not a Fit") else 1, rec["machine"].get("company", ""))
    rows = sorted(rows, key=sort_key)
    for i, rec in enumerate(rows):
        r = first + i
        m, h = rec["machine"], rec["human"]
        disp = _review_status(rec)
        fit = m.get("fit_score", "")
        fit_disp = "❌" if m.get("fit_disqualifier") not in (None, "", "none") else fit
        vals = [rec["_cid"], disp, m.get("company", ""), m.get("role", ""),
                m.get("lane", ""), m.get("location", ""), m.get("cycle", ""),
                fit_disp, m.get("fit_why", ""), ("Apply ↗" if m.get("url") else ""),
                h.get("notes", "")]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(j in (4, 9, 11)))
        sf = STATUS_FILL.get(disp)
        if sf:
            sc = ws.cell(row=r, column=2)
            sc.fill = PatternFill("solid", fgColor=sf[0])
            sc.font = Font(bold=True, size=10, color=sf[1])
        if m.get("url"):
            ac = ws.cell(row=r, column=10)
            ac.hyperlink = m["url"]
            ac.font = Font(size=10, color="2563EB", underline="single")
            ac.alignment = Alignment(horizontal="center", vertical="center")
    last = first + len(rows) - 1 if rows else first
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"B{hdr}:{get_column_letter(len(REVIEW_HEADERS))}{last}"
    if rows:
        _add_dv(ws, "B", STATUS_OPTS, first, last)
        for val, (bg, fg) in STATUS_FILL.items():
            ws.conditional_formatting.add(f"B{first}:B{last}",
                CellIsRule(operator="equal", formula=[f'"{val}"'],
                           fill=PatternFill("solid", fgColor=bg),
                           font=Font(color=fg, bold=True, size=10)))
    return first, last


# ── public: write the whole board ─────────────────────────────────────────────
def write_board(store, path: str | Path, generated_at: str) -> dict:
    """Regenerate the three-sheet board from the store. Atomic save. Returns counts.
    Rows are routed by classify_row(): live 'To Apply' -> Curated Queue; real application
    statuses -> My Applications; Skip/Not-a-Fit OR stale-with-a-note -> Reviewed/Skipped
    (so a note you wrote is never silently lost); stale-with-no-note -> dropped (kept in JSON)."""
    check_lock(path)

    queue_rows, app_rows, reviewed_rows = [], [], []
    bucket = {"queue": queue_rows, "application": app_rows, "reviewed": reviewed_rows}
    for cid, rec in store.items():
        rec = {"_cid": cid, "machine": rec.get("machine", {}), "human": rec.get("human", {})}
        dest = classify_row(rec)
        if dest in bucket:
            bucket[dest].append(rec)
        # dest == "drop" -> off the board (still in JSON)

    wb = Workbook()
    q = wb.active
    q.title = "Curated Queue"
    _build_queue(q, queue_rows)
    a = wb.create_sheet("My Applications")
    app_first, app_last = _build_apps(a, app_rows)
    rv = wb.create_sheet("Reviewed")
    _build_reviewed(rv, reviewed_rows)
    _build_summary(wb, app_first, app_last, len(queue_rows), len(reviewed_rows), generated_at)
    meta = wb.create_sheet("_meta")
    meta.sheet_state = "hidden"
    meta["A1"] = "schema_version"; meta["B1"] = SCHEMA_VERSION
    meta["A2"] = "generated_at"; meta["B2"] = generated_at

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=str(p.parent), suffix=".xlsx.tmp")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, p)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    return {"queue": len(queue_rows), "applications": len(app_rows),
            "reviewed": len(reviewed_rows)}

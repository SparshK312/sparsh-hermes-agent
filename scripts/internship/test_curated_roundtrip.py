#!/usr/bin/env python3
"""
Verification suite for the living curated board. The critical test is round-trip
preservation: manual edits must survive a full regenerate + re-rank, with all
formatting (data-validation / conditional-formatting / hyperlinks) intact.

Runs against a throwaway copy of the real store in a temp dir — never touches the
live board. Exit 0 = all pass.
"""
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
VAULT = Path(os.environ.get("HERMES_VAULT")
             or ("/home/hermes/vault" if Path("/home/hermes/vault").exists()
                 else str(Path.home() / "Documents" / "School Vault - UofT")))
sys.path.insert(0, str(VAULT / "Scripts"))

from openpyxl import load_workbook
from build_curated_xlsx import check_lock, read_back_human, write_board, ID_HEADER
from curated_store import CuratedStore

STORE_PATH = VAULT / "06 - Internships" / "Internship Pipeline" / "curated_postings.json"

PASS = 0
FAIL = 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✓ {name}")
    else:
        FAIL += 1
        print(f"  ✗ FAIL: {name}")


def _find_sheet_cols(ws):
    """Return (header_row_idx, {header_lower: col_idx_1based}) for a data sheet."""
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if ID_HEADER in cells:
            return ri, {h.lower(): i + 1 for i, h in enumerate(cells)}
    return None, {}


def _edit_xlsx(path, cid, **edits):
    """Simulate the user editing a row in Excel: find the row by _id, set columns."""
    wb = load_workbook(path)
    done = False
    for ws in wb.worksheets:
        if ws.title.startswith("_") or ws.title == "Summary":
            continue
        hr, cols = _find_sheet_cols(ws)
        if hr is None:
            continue
        idc = cols.get(ID_HEADER.lower())
        for r in range(hr + 1, ws.max_row + 1):
            if str(ws.cell(row=r, column=idc).value or "") == cid:
                for header, val in edits.items():
                    c = cols.get(header.lower())
                    if c:
                        ws.cell(row=r, column=c, value=val)
                done = True
                break
        if done:
            break
    wb.save(path)
    return done


def main():
    tmp = Path(tempfile.mkdtemp())
    xlsx = tmp / "Curated Board.xlsx"
    gen = "2026-06-20 12:00"

    store = CuratedStore(STORE_PATH).load()
    queue_cids = [cid for cid, e in store.items()
                  if (e.get("human", {}).get("status") or "") in ("", "To Apply")
                  and not e.get("machine", {}).get("dead")]
    app_cids = [cid for cid, e in store.items()
                if (e.get("human", {}).get("status") or "") not in ("", "To Apply")]
    assert len(queue_cids) >= 3, "need >=3 queue rows to test"
    assert len(app_cids) >= 1, "need >=1 application row to test Applied date"
    cid_status, cid_prio, cid_notes = queue_cids[:3]
    cid_app = app_cids[0]   # an actioned row living in My Applications (has Applied col)

    print("1. baseline write")
    write_board(store, xlsx, gen)
    check("xlsx created", xlsx.exists())

    print("2. simulate user edits in Excel")
    _edit_xlsx(xlsx, cid_status, Status="Applied")     # queue: set status (date goes in MyApps)
    _edit_xlsx(xlsx, cid_prio, Priority="P0")
    _edit_xlsx(xlsx, cid_notes, Notes="ROUNDTRIP TEST NOTE")
    _edit_xlsx(xlsx, cid_app, Applied="2026-06-20")    # My Applications: set the applied date

    print("3. read-back picks up the edits")
    back = read_back_human(xlsx)
    check("status edit read back", back.get(cid_status, {}).get("status") == "Applied")
    check("applied date read back (My Applications)",
          back.get(cid_app, {}).get("applied_date") == "2026-06-20")
    check("priority override read back", back.get(cid_prio, {}).get("priority_override") == "P0")
    check("notes read back", back.get(cid_notes, {}).get("notes") == "ROUNDTRIP TEST NOTE")

    print("4. merge + regenerate (simulate a refresh; rows re-sort)")
    for cid, human in back.items():
        store.set_human(cid, human)
    # simulate a day passing: nudge hotness so the queue re-sorts
    for cid, e in list(store.items())[:20]:
        m = e.get("machine", {})
        if "hotness" in m:
            m["hotness"] = max(0, int(m["hotness"]) - 7)
    write_board(store, xlsx, "2026-06-21 09:00")

    print("5. edits survived the regenerate + re-rank")
    back2 = read_back_human(xlsx)
    check("status survived", back2.get(cid_status, {}).get("status") == "Applied")
    check("applied survived", back2.get(cid_app, {}).get("applied_date") == "2026-06-20")
    check("priority survived", back2.get(cid_prio, {}).get("priority_override") == "P0")
    check("notes survived", back2.get(cid_notes, {}).get("notes") == "ROUNDTRIP TEST NOTE")

    print("6. actioned row moved Queue -> My Applications")
    wb = load_workbook(xlsx)
    q_ids = _ids_in_sheet(wb["Curated Queue"])
    a_ids = _ids_in_sheet(wb["My Applications"])
    check("Applied row left the queue", cid_status not in q_ids)
    check("Applied row in My Applications", cid_status in a_ids)
    check("P0 row still in queue", cid_prio in q_ids)

    print("7. formatting intact in regenerated file")
    q = wb["Curated Queue"]
    check("data validations present", len(list(q.data_validations.dataValidation)) >= 3)
    check("conditional formatting present", len(list(q.conditional_formatting)) >= 1)
    # find the Apply column by header (robust to column reordering)
    _hr, _cols = _find_sheet_cols(q)
    apply_col = _cols.get("apply")
    has_link = apply_col and any(q.cell(row=r, column=apply_col).hyperlink
                                 for r in range(4, min(q.max_row, 40) + 1))
    check("hyperlinks present (Apply col)", bool(has_link))
    check("_meta hidden sheet present", "_meta" in wb.sheetnames)

    print("8. dead-link handling")
    store2 = CuratedStore(STORE_PATH).load()
    dead_unactioned = next(cid for cid, e in store2.items()
                           if (e.get("human", {}).get("status") or "") in ("", "To Apply"))
    store2.entry(dead_unactioned)["machine"]["dead"] = True
    # a dead row the user already applied to (any actioned row)
    applied_dead = next(cid for cid, e in store2.items()
                        if (e.get("human", {}).get("status") or "") not in ("", "To Apply"))
    store2.entry(applied_dead)["machine"]["dead"] = True
    store2.set_human(applied_dead, {"status": "Applied"})
    write_board(store2, xlsx, gen)
    wb2 = load_workbook(xlsx)
    check("dead+unactioned dropped from queue",
          dead_unactioned not in _ids_in_sheet(wb2["Curated Queue"]))
    check("dead+applied kept in My Applications",
          applied_dead in _ids_in_sheet(wb2["My Applications"]))

    print("9. lock detection")
    lock = xlsx.with_name("~$" + xlsx.name)
    lock.write_text("x")
    locked = False
    try:
        check_lock(xlsx)
    except RuntimeError:
        locked = True
    check("aborts when ~$ lock present", locked)
    lock.unlink()

    print("10. orphan (hand-added row) preserved")
    store3 = CuratedStore(STORE_PATH).load()
    store3.add_orphan("hist/manual/my-dream-job", {"status": "Networking", "notes": "cold email sent"})
    check("orphan stored", store3.get("hist/manual/my-dream-job") is not None)
    check("orphan has human data",
          store3.get("hist/manual/my-dream-job")["human"]["status"] == "Networking")

    shutil.rmtree(tmp, ignore_errors=True)
    print(f"\n{PASS} passed, {FAIL} failed")
    return 1 if FAIL else 0


def _ids_in_sheet(ws):
    hr, cols = _find_sheet_cols(ws)
    if hr is None:
        return set()
    idc = cols.get(ID_HEADER.lower())
    return {str(ws.cell(row=r, column=idc).value) for r in range(hr + 1, ws.max_row + 1)
            if ws.cell(row=r, column=idc).value}


if __name__ == "__main__":
    sys.exit(main())
